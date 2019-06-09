# -*- coding: utf-8 -*-
# !/usr/bin/python
import os
from multiprocessing.dummy import Pool

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


class ExcelOperations:
    wb = Workbook()  # wb作为一个类对象被使用,在方法内部使用会造成比如创建sheet覆盖上个sheet的问题.

    def __int__(self):
        pass

    def dir_files(self, dir_name, file_ext):
        """
        获取目录下的扩展名为file_ext的文件列表
        :param dir_name: 目录
        :param file_ext: 扩展名，不带点(.)
        :return:
        """
        file_arr = []
        t = os.walk(dir_name, topdown=True, onerror=None, followlinks=False)
        for root, dirs, files in t:
            for file in files:
                ext = file.split('.')[-1]
                if ext == file_ext and file.find('~$') != 0:
                    file_arr.append(root + '\\' + file)
                    print("根目录：" + root + '，文件夹列表：' + str(dirs) + '，文件：' + str(file))
        return file_arr

    def get_data(self, filename):
        """
        获取文件中的数据
        :param filename: 需要获取数据的文件
        :return values: 获取到的数据[{'sheet_name': sheet.title, 'sheet_data': []}, ...]
        """
        values = []
        wb = load_workbook(filename)
        for index, sheet in enumerate(wb):
            values.append({'sheet_name': sheet.title, 'sheet_data': []})
            for row in sheet.iter_rows():
                row_data = []
                for cell in row:
                    row_data.append(cell.value)
                values[index]['sheet_data'].append(row_data)
        return values

    def set_data(self, sheet_name, data, filename, header_data):
        sheet = self.wb.create_sheet(sheet_name, 0)
        sheet.append([x[0] for x in header_data])
        row = sheet.row_dimensions[1]
        for i, item in enumerate(header_data):
            sheet.column_dimensions[get_column_letter(i + 1)].width = item[1]
        row.font = Font(bold=True)
        for i, x in enumerate(data):
            sheet.append(x)

        self.wb.save(filename)

    def test(self):
        """
        创建sheet
        :return:
        """
        import datetime
        sheet1 = self.wb.active

        sheet1.title = 'DangDang Book'  # 添加sheet
        sheet2 = self.wb.create_sheet('Amazon Book')
        sheet3 = self.wb.create_sheet('JD Book')

        sheet1['A1'] = 'A1'  # 写入第1行第A列数据为A1
        sheet1['B1'] = 'B1'
        sheet2['B1'] = 'B1'
        sheet1.append([1, 2, 3])
        sheet1['A2'] = datetime.datetime.now()

        self.wb.save('example.xlsx')  # 保存到example.xlsx


if __name__ == '__main__':
    p = Pool()
    e = ExcelOperations()
    excels_dir = 'E:\\ProgramData\\GitHub\\PyLang\\Crawlers\\BuyBook\\static_files\\excels'

    print('父进程pid：%s.' % os.getpid())

    p.apply_async(e.get_data, args=("../../BuyBook/static_files/excels/xlsx/详细拉卡拉POS机领用记录3月20日（最新）.xlsx",))
    files_arr = p.apply_async(e.dir_files, args=(excels_dir, 'xls'))
    p.apply_async(e.xlstoxlsx, args=(files_arr.get(),))

    p.close()
    p.join()
